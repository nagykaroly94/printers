import asyncio
from openpyxl import Workbook
import io
from flask import Flask, Response, jsonify
from pysnmp.hlapi.v3arch.asyncio import (
    SnmpEngine, CommunityData, UdpTransportTarget, ContextData,
    ObjectType, ObjectIdentity, get_cmd
)
import threading
from flask import render_template, request
import mysql.connector
import configparser
import os

app = Flask(__name__)

results = {}
running = False

def get_db_connection():
    # Absolút útvonal a config.ini-hez
    base_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(base_dir, "config.ini")

    config = configparser.ConfigParser()
    config.read(config_path)

    db = mysql.connector.connect(
        host=config["mysql"]["host"],
        user=config["mysql"]["user"],
        password=config["mysql"]["password"],
        database=config["mysql"]["database"]
    )

    return db

# Nyomtatók lekérdezése MySQL-ből
def load_printers():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT azonosito, gep_helye, ip, tipus, gyari_szam, uzemelteto, cim
        FROM nyomtatok
    """)
    rows = cursor.fetchall()

    printers = {}
    for r in rows:
        printers[r["ip"]] = r  # itt már az egész sor JSON-szerű objektumként

    return printers

ip_locations = load_printers()
# SNMP lekérdezés
async def snmp_get(ip, oid):
    transport = await UdpTransportTarget.create((ip,161), timeout=2, retries=1)
    errorIndication, errorStatus, errorIndex, varBinds = await get_cmd(
        SnmpEngine(),
        CommunityData('public', mpModel=0),
        transport,
        ContextData(),
        ObjectType(ObjectIdentity(oid))
    )
    if errorIndication or errorStatus:
        return None
    for varBind in varBinds:
        return varBind[1]

async def get_printer_data(ip):
    """Lekérdezi a nyomtató típusát, oldalszámát és sorozatszámát"""

    type_result = await snmp_get(ip, '1.3.6.1.2.1.1.1.0')
    if type_result is None:
        return (
            "Nem sikerült a lekérdezés",
            "Nem sikerült a lekérdezés",
            "Nem sikerült a lekérdezés"
        )

    full_type = type_result.prettyPrint()
    if not full_type.strip():
        printer_type = "Nem sikerült a lekérdezés"
    else:
        printer_type = " ".join(full_type.split()[:3])

    # alap OID-ok
    page_oid = '1.3.6.1.2.1.43.10.2.1.4.1.1'
    type_oid = None
    serial_oid = '1.3.6.1.2.1.43.5.1.1.17.1'

    # Canon imageRUNNER1133 kivétel
    if "Canon imageRUNNER1133" in full_type:
        page_oid = '1.3.6.1.4.1.1602.1.11.1.3.1.4.113'

    # Canon iR-ADV 525 III kivétel
    if "Canon iR-ADV 525 III" in full_type:
        page_oid = '1.3.6.1.2.1.43.10.2.1.4.1.1'
        type_oid = '1.3.6.1.2.1.25.3.2.1.3.1'
        serial_oid = '1.3.6.1.2.1.43.5.1.1.17.1'

    # típus lekérdezés speciális OID-dal ha kell
    if type_oid:
        type_result = await snmp_get(ip, type_oid)
        if type_result:
            full_type = type_result.prettyPrint()
            if full_type.strip():
                printer_type = full_type
            else:
                printer_type = "Nem sikerült a lekérdezés"

    # oldalszám lekérdezés
    page_result = await snmp_get(ip, page_oid)
    pages = page_result if page_result is not None else "Nem sikerült a lekérdezés"

    # sorozatszám lekérdezés
    serial_result = await snmp_get(ip, serial_oid)
    serial = serial_result.prettyPrint() if serial_result is not None else "Nem sikerült a lekérdezés"

    return printer_type, pages, serial
# Lekérdezés futtatása
async def run_query():
    global running, results
    ip_locations = load_printers()
    running = True
    results = {}

    for ip, r in ip_locations.items():
        printer_id = r["azonosito"]
        location = r["gep_helye"]
        uzemelteto = r.get("uzemelteto", "")
        cim = r.get("cim", "")
        try:
            printer_type, page_count, serial = await get_printer_data(ip)

            # --- JSON kompatibilis ---
            try:
                page_count_int = int(page_count)
                page_count_display = page_count_int
            except:
                page_count_int = None
                page_count_display = "Nem sikerült a lekérdezés"

            if serial is None:
                serial_display = "Nem sikerült a lekérdezés"
            else:
                serial_display = str(serial)

            # --- results MINDIG töltődik ---
            results[ip] = {
                "id": printer_id,
                "name": location,
                "ip": ip,
                "type": printer_type,
                "serial": serial_display,
                "pages": page_count_display,
                "cim": cim,
                "uzemelteto": uzemelteto
            }

            # --- DB LOGIKA KÜLÖN ---
            if (
                printer_type != "Nem sikerült a lekérdezés" and
                serial_display != "Nem sikerült a lekérdezés" and
                printer_type is not None and
                serial_display is not None and
                page_count_int is not None
            ):
                db = get_db_connection()
                cursor = db.cursor(dictionary=True)

                cursor.execute("""
                    SELECT tipus, gyari_szam, oldalszam 
                    FROM nyomtatok 
                    WHERE azonosito=%s
                """, (printer_id,))

                row = cursor.fetchone()

                if row:
                    if not (
                        row["tipus"] == printer_type and
                        row["gyari_szam"] == serial_display and
                        row["oldalszam"] == page_count_int
                    ):
                        cursor.execute("""
                            UPDATE nyomtatok 
                            SET tipus=%s, gyari_szam=%s, oldalszam=%s
                            WHERE azonosito=%s
                        """, (
                            printer_type,
                            serial_display,
                            page_count_int,
                            printer_id
                        ))

        except Exception:
            results[ip] = {
                "id": printer_id,
                "name": location,
                "type": "Nem sikerült a lekérdezés",
                "pages": "Nem sikerült a lekérdezés",
                "serial": "Nem sikerült a lekérdezés",
                "cim": cim,
                "uzemelteto": uzemelteto,
                "ip": ip
            }

    db.commit()
    running = False
# Flask útvonalak
def save_monthly_snapshot():
    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        INSERT INTO nyomtato_havi_allas (nyomtato_id, uzemelteto, cim, datum, oldalszam)
        SELECT azonosito, uzemelteto, cim, DATE_FORMAT(CURDATE(), '%Y-%m-01'), oldalszam
        FROM nyomtatok
        ON DUPLICATE KEY UPDATE oldalszam = VALUES(oldalszam), rogzitve = NOW(), uzemelteto = VALUES(uzemelteto), cim = VALUES(cim);
    """)

    db.commit()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/start")
def start():
    global running
    if not running:
        thread = threading.Thread(target=lambda: asyncio.run(run_query()))
        thread.start()
    return "ok"

@app.route("/status")
def status():
    return jsonify({"results": results, "total": len(ip_locations), "running": running})

@app.route("/printer_pages/<path:table_name>.xlsx")
def download_table_xlsx(table_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"

    ws.append(["Azonosító","Hely","IP cím","Típus","Sorozatszám","Oldalszám"])

    for r in results.values():
        company = r.get("uzemelteto", "Nincs üzemeltető")
        address = r.get("cim", "Nincs cím")
        key = f"{company} - {address}"

        if key == table_name:
            ws.append([
                r["id"],
                r["name"],
                r["ip"],
                r["type"],
                r["serial"],
                r["pages"]
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": f"attachment; filename={table_name}.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

    return Response(output, headers=headers)

@app.route("/printer_pages.xlsx")
def download_all_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"
    ws.append(["Azonosító","Hely","IP cím","Típus","Sorozatszám","Oldalszám","Üzemeltető","Cím"])
    for r in results.values():
        ws.append([r["id"], r["name"], r["ip"], r["type"], r["serial"], r["pages"], r["uzemelteto"], r["cim"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=Teljes_tablazat.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "X-Content-Type-Options": "nosniff"  # Chrome biztonságosabbnak látja
    }

    return Response(output, headers=headers)

@app.route("/add_printer", methods=["POST"])
def add_printer():
    try:
        azonosito = request.form.get("azonosito")
        gep_helye = request.form.get("gep_helye")
        ip = request.form.get("ip")
        tipus = request.form.get("tipus")
        gyari_szam = request.form.get("gyari_szam")
        uzemelteto = request.form.get("uzemelteto")
        cim = request.form.get("cim")

        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute("""
            INSERT INTO nyomtatok 
            (azonosito, gep_helye, ip, tipus, gyari_szam, uzemelteto, cim)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            azonosito,
            gep_helye,
            ip,
            tipus,
            gyari_szam,
            uzemelteto,
            cim
        ))

        db.commit()

        return {"success": True}

    except Exception as e:
        return {"success": False, "error": str(e)}

@app.route("/get_printers")
def get_printers():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM nyomtatok")
    return jsonify(cursor.fetchall())

@app.route("/delete_printer/<azonosito>", methods=["DELETE"])
def delete_printer(azonosito):
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("DELETE FROM nyomtatok WHERE azonosito=%s", (azonosito,))
    db.commit()
    return {"success": True}

@app.route("/update_printer", methods=["POST"])
def update_printer():
    data = request.json

    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("""
        UPDATE nyomtatok SET
            gep_helye=%s,
            ip=%s,
            tipus=%s,
            gyari_szam=%s,
            uzemelteto=%s,
            cim=%s
        WHERE azonosito=%s
    """, (
        data["gep_helye"],
        data["ip"],
        data["tipus"],
        data["gyari_szam"],
        data["uzemelteto"],
        data["cim"],
        data["azonosito"]
    ))

    db.commit()
    return {"success": True}

@app.route("/save_monthly", methods=["POST"])
def save_monthly():
    try:
        save_monthly_snapshot()
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.route("/get_full_diff")
def getfulldiff():
    honapok = {
        "January": "Január",
        "February": "Február",
        "March": "Március",
        "April": "Április",
        "May": "Május",
        "June": "Június",
        "July": "Július",
        "August": "Augusztus",
        "September": "Szeptember",
        "October": "Október",
        "November": "November",
        "December": "December"
    }
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    

    cursor.execute("""
        SELECT 
            curr.nyomtato_id, 
            curr.uzemelteto, 
            curr.cim, 
            curr.honap_nev, 
            curr.oldalszam AS aktualis, 
            prev.oldalszam AS elozo, 
            (curr.oldalszam - prev.oldalszam) AS kulonbseg 
        FROM 
            nyomtato_havi_allas curr 
            LEFT JOIN nyomtato_havi_allas prev ON curr.nyomtato_id = prev.nyomtato_id 
            AND prev.datum = DATE_SUB(curr.datum, INTERVAL 1 MONTH) 
        WHERE 
            curr.honap_nev LIKE MONTHNAME(
                CURDATE()
            ) 
        ORDER BY 
            `uzemelteto`, 
            `cim`;
    """)
    rows = cursor.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"
    ws.append(["Nyomtató ID","Üzemeltető","Cím","Aktuális hónap","Aktuális nyomatszám","Előző nyomatszám","Különbség"])
    for r in rows:
        ws.append([r["nyomtato_id"], r["uzemelteto"], r["cim"], honapok.get(r["honap_nev"]), r["aktualis"], r["elozo"], r["kulonbseg"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=Tárgy havi nyomatszámok.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "X-Content-Type-Options": "nosniff"  # Chrome biztonságosabbnak látja
    }

    return Response(output, headers=headers)

if __name__=="__main__":
    app.run(host="0.0.0.0", port=5000)