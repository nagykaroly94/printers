import asyncio
import threading
import io
import os
import configparser
import mysql.connector
import ipaddress
import bcrypt

from flask import Flask, Response, jsonify, render_template, request, redirect, url_for
from openpyxl import Workbook
from pysnmp.hlapi.v3arch.asyncio import (
    SnmpEngine, CommunityData, UdpTransportTarget, ContextData,
    ObjectType, ObjectIdentity, get_cmd
)
from datetime import datetime, timedelta

from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    logout_user,
    login_required,
    current_user
)

config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.ini")
config = configparser.ConfigParser()
config.read(config_path)

app = Flask(__name__)
app.config["REMEMBER_COOKIE_DURATION"] = timedelta(days=40)  # 40 napos cookie élettartam
app.config["SECRET_KEY"] = config["app"]["secret"]

login_manager = LoginManager()
login_manager.init_app(app)

login_manager.login_view = "login"
login_manager.login_message = "A folytatáshoz be kell jelentkezned."
login_manager.login_message_category = "error"

# -------------------------
# THREAD SAFE STATE
# -------------------------
results = {}
running = False
results_lock = threading.Lock()
ip_locations = {}
TOTAL_PRINTERS = 0
processed = 0
processed_lock = threading.Lock()
live_updates = []
live_updates_lock = threading.Lock()

class User(UserMixin):

    def __init__(self, username, password_hash):
        self.id = username
        self.username = username
        self.password_hash = password_hash

def invalidate_cache():
    global results
    with results_lock:
        results = {}

# -------------------------
# DB CONNECTION
# -------------------------
def get_db_connection():
    global config
    return mysql.connector.connect(
        host=config["mysql"]["host"],
        user=config["mysql"]["user"],
        password=config["mysql"]["password"],
        database=config["mysql"]["database"]
    )

def init_app_state():
    global results
    results = load_initial_state()

def load_initial_state():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM nyomtatok")
    rows = cursor.fetchall()
    cursor.close()
    db.close()

    return {
        r["azonosito"]: {
            "id": r["azonosito"],
            "name": r.get("gep_helye") or "N/A",
            "ip": r.get("ip") or "N/A",
            "type": r.get("tipus") or "N/A",
            "serial": r.get("gyari_szam") or "N/A",
            "pages": r.get("oldalszam") or "N/A",
            "cim": r.get("cim") or "N/A",
            "uzemelteto": r.get("uzemelteto") or "N/A",
            "tablazat": r.get("tablazat"),
            "status": r.get("status") or "idle",
            "rogzitve": (r["updated_at"].strftime("%Y-%m-%d %H:%M:%S") if isinstance(r.get("updated_at"), datetime) else r.get("updated_at")),
        }
        for r in rows
    }

# -------------------------
# LOAD PRINTERS
# -------------------------
def load_printers():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT * FROM nyomtatok")
    rows = cursor.fetchall()

    cursor.close()
    db.close()

    global TOTAL_PRINTERS
    TOTAL_PRINTERS = len(rows)

    return {r["azonosito"]: r for r in rows}


# -------------------------
# SNMP GET
# -------------------------
async def snmp_get(ip, oid):
    try:
        transport = await UdpTransportTarget.create((ip, 161), timeout=5, retries=2)

        errorIndication, errorStatus, errorIndex, varBinds = await get_cmd(
            SnmpEngine(),
            CommunityData('public', mpModel=0),
            transport,
            ContextData(),
            ObjectType(ObjectIdentity(oid))
        )

        if errorIndication or errorStatus:
            return None

        for v in varBinds:
            return v[1]

    except Exception:
        return None


# -------------------------
# PRINTER DATA
# -------------------------
async def get_printer_data(ip):
    type_result = await snmp_get(ip, '1.3.6.1.2.1.1.1.0')

    if type_result is None:
        return None, None, None

    full_type = type_result.prettyPrint() or ""
    full_type_upper = full_type.upper()

    printer_type = " ".join(full_type.split()[:3]) if full_type else None

    # -------------------------
    # DEFAULT OID-ek
    # -------------------------
    page_oid = '1.3.6.1.2.1.43.10.2.1.4.1.1'

    serial_oids = [
        '1.3.6.1.2.1.43.5.1.1.17.1',  # standard Printer-MIB serial
    ]

    # -------------------------
    # HP FIX
    # -------------------------
    hp_type = None

    if "HP" in full_type_upper and (
        "JETDIRECT" in full_type_upper or
        "ETHERNET" in full_type_upper or
        "MULTI-ENVIRONMENT" in full_type_upper
    ):

        alt = await snmp_get(ip, '1.3.6.1.2.1.25.3.2.1.3.1')

        if alt:
            alt_str = alt.prettyPrint()

            if alt_str and alt_str.strip():
                hp_type = alt_str.strip()

    if hp_type:
        printer_type = hp_type

    # -------------------------
    # CANON FIX
    # -------------------------
    if "CANON" in full_type_upper:

        # imageRUNNER 1133 family
        if "1133" in full_type_upper:
            page_oid = '1.3.6.1.4.1.1602.1.11.1.3.1.4.113'

        # iR-ADV
        elif "IR-ADV" in full_type_upper:
            page_oid = '1.3.6.1.2.1.43.10.2.1.4.1.1'

        # Canon serial fallback OID-ek
        serial_oids.extend([

            # VALÓDI serial LBP6650/P
            '1.3.6.1.4.1.1602.1.2.1.4.0',

            # engine/controller ID
            '1.3.6.1.4.1.1602.1.3.1.1.1.1.1',

            # egyéb Canon
            '1.3.6.1.4.1.1602.1.11.1.1.3.1.1',

            # model
            '1.3.6.1.4.1.1602.1.1.1.2.0'
        ])

    # -------------------------
    # PAGE COUNT
    # -------------------------
    page_result = await snmp_get(ip, page_oid)

    pages = None

    if page_result is not None:
        try:
            pages = int(page_result)
        except Exception:
            pages = None

    # -------------------------
    # SERIAL DETECT
    # -------------------------
    serial = None

    for oid in serial_oids:

        try:
            result = await snmp_get(ip, oid)

            if result:
                value = result.prettyPrint().strip()

                print(f"{ip} SERIAL OID {oid} => {value}")

                if (
                    value and
                    value.upper() not in ["NONE", "N/A"] and
                    "NOSUCH" not in value.upper() and
                    value.upper() not in [
                        "SN-E2",
                        "LBP6650"
                    ]
                ):
                    serial = value
                    break

        except Exception as e:
            print(f"{ip} SERIAL ERROR {oid}: {e}")

    return printer_type, pages, serial

# -------------------------
# MAIN QUERY
# -------------------------
async def run_query():
    global running, results, processed

    running = True

    with processed_lock:
        processed = 0

    local_results = {}
    printer_locations = load_printers()

    semaphore = asyncio.Semaphore(20)

    async def get_data(printer_id, r):
        async with semaphore:

            db = get_db_connection()
            cursor = db.cursor()

            try:
                ip = r["ip"]
                try: # Do we even need the snmp, is the IP Addr. valid?
                    ipaddress.ip_address(ip)
                except ValueError:
                    raise ValueError(f"Invalid IP address: {ip}")

                printer_type, page_count, serial = await get_printer_data(ip)

                # -------------------------
                # SNMP SUCCESS CHECK
                # -------------------------
                success = (
                    printer_type is not None or
                    page_count is not None or
                    serial is not None
                )

                # -------------------------
                # PAGE COUNT LOGIKA
                # -------------------------
                try:
                    page_count_int = int(page_count) if page_count not in [None, "", "N/A"] else int(r.get("oldalszam") or 0)
                except:
                    page_count_int = None

                old_pages = int(r.get("oldalszam") or 0)
                pages_changed = (
                    page_count_int is not None and
                    old_pages is not None and
                    page_count_int != old_pages
                )

                # -------------------------
                # FINAL OUTPUT (FRONTEND)
                # -------------------------
                final_type = printer_type or r.get("tipus")
                final_serial = serial or r.get("gyari_szam")

                data = {
                    "id": r["azonosito"],
                    "name": r.get("gep_helye") or "N/A",
                    "ip": r.get("ip") or "N/A",
                    "type": final_type or "N/A",
                    "serial": final_serial or "N/A",
                    "pages": page_count_int or "N/A",
                    "cim": r.get("cim") or "N/A",
                    "uzemelteto": r.get("uzemelteto") or "N/A",
                    "tablazat": r.get("tablazat"),
                    "status": "ok" if success else "error",
                    "rogzitve": (r.get("updated_at").isoformat() if isinstance(r.get("updated_at"), datetime) else r.get("updated_at"))
                }

                local_results[r["azonosito"]] = data

                with live_updates_lock:
                    live_updates.append(data)

                # -------------------------
                # DB UPDATE LOGIKA
                # -------------------------
                updates = []
                values = []

                # csak ha jött SNMP adat
                if printer_type:
                    updates.append("tipus=%s")
                    values.append(printer_type)

                if serial:
                    updates.append("gyari_szam=%s")
                    values.append(serial)

                if pages_changed:
                    updates.append("oldalszam=%s")
                    values.append(page_count_int)

                if success:
                    updates.append("updated_at = %s")
                    values.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                # csak akkor írunk DB-be, ha van mit frissíteni
                if updates:
                    values.append(r["azonosito"])

                    cursor.execute(f"""
                        UPDATE nyomtatok
                        SET {", ".join(updates)}
                        WHERE azonosito=%s
                    """, values)

                    db.commit()

            except Exception as e:
                local_results[r["azonosito"]] = {
                    "id": r["azonosito"],
                    "name": r.get("gep_helye") or "N/A",
                    "ip": r.get("ip") or "N/A",
                    "type": r.get("tipus") or "N/A",
                    "serial": r.get("gyari_szam") or "N/A",
                    "pages": r.get("oldalszam") or "N/A",
                    "cim": r.get("cim") or "N/A",
                    "uzemelteto": r.get("uzemelteto") or "N/A",
                    "tablazat": r.get("tablazat"),
                    "status": "error",
                    "rogzitve": (r.get("updated_at").isoformat() if isinstance(r.get("updated_at"), datetime) else r.get("updated_at"))
                }

            finally:
                cursor.close()
                db.close()

                global processed
                with processed_lock:
                    processed = processed + 1

    await asyncio.gather(*(get_data(id_, r) for id_, r in printer_locations.items()))

    with results_lock:
        results = local_results

    running = False

# -------------------------
# SNAPSHOT
# -------------------------
def save_monthly_snapshot():
    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        INSERT INTO nyomtato_havi_allas (nyomtato_id, uzemelteto, cim, datum, oldalszam)
        SELECT azonosito, uzemelteto, cim, DATE_FORMAT(CURDATE(), '%Y-%m-01'), oldalszam
        FROM nyomtatok
        ON DUPLICATE KEY UPDATE
            oldalszam = VALUES(oldalszam),
            rogzitve = NOW(),
            uzemelteto = VALUES(uzemelteto),
            cim = VALUES(cim);
    """)

    db.commit()
    cursor.close()
    db.close()


# -------------------------
# FLASK ROUTES
# -------------------------
@app.before_request
def check_login():

    # Login oldal szabadon elérhető
    if request.endpoint == "login":
        return

    # CSS, JS, képek stb. szabadon elérhetők
    if request.endpoint == "static":
        return

    # Minden más csak bejelentkezve
    if not current_user.is_authenticated:
        return redirect(url_for("login"))


@app.route("/", methods=["GET", "POST"])
def login():

    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":

        username = request.form.get("username")
        password = request.form.get("password")

        print("LOGIN POST:", username)

        db = None
        cursor = None

        try:
            db = get_db_connection()
            cursor = db.cursor()

            cursor.execute(
                """
                SELECT felhasznalonev, jelszo
                FROM felhasznalok
                WHERE felhasznalonev = %s
                """,
                (username,)
            )

            result = cursor.fetchone()

            print("DB RESULT:", result)

            if result is not None:

                db_username = result[0]
                stored_password = result[1]

                if bcrypt.checkpw(
                    password.encode("utf-8"),
                    stored_password.encode("utf-8")
                ):

                    user = User(
                        username=db_username,
                        password_hash=stored_password
                    )

                    remember = request.form.get("remember") == "on"

                    login_user(user, remember=remember)

                    print("SIKERES LOGIN:", current_user.username)

                    return redirect(url_for("index"))

            return render_template(
                "login.html",
                hiba="Helytelen felhasználónév vagy jelszó."
            )

        except Exception as e:

            print("LOGIN HIBA:", e)

            return render_template(
                "login.html",
                hiba="Hiba történt a bejelentkezés során."
            )

        finally:

            if cursor:
                cursor.close()

            if db:
                db.close()

    return render_template("login.html")

@app.route("/start")
def start():
    global running
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if running:
        return f"already running"

    thread = threading.Thread(target=lambda: asyncio.run(run_query()))
    thread.start()

    return f"ok"

@app.route("/status")
def status():
    global live_updates

    with processed_lock:
        prog = processed

    with live_updates_lock:
        updates = live_updates.copy()
        live_updates.clear()

    return jsonify({
        "processed": prog,
        "total": TOTAL_PRINTERS,
        "running": running,
        "updates": updates
    })

@app.route("/printer_pages/<path:table_name>.xlsx")
def download_table_xlsx(table_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"

    ws.append(["Azonosító","Hely","IP cím","Típus","Sorozatszám","Oldalszám"])

    for r in results.values():
        company = r.get("uzemelteto", "Nincs üzemeltető")
        address = r.get("cim", "Nincs cím")

        if table_name == r.get("tablazat", ""):
            ws.append([
                r["id"],
                r["name"],
                r["ip"],
                r["type"],
                r["serial"],
                r["pages"]
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": f"attachment; filename={table_name}.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

    return Response(output, headers=headers)

# -------------------------
# EXCEL EXPORT (ALL)
# -------------------------
@app.route("/printer_pages.xlsx")
def download_all_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"

    ws.append(["Azonosító","Hely","IP","Típus","Sorozatszám","Oldalszám","Üzemeltető","Cím"])

    with results_lock:
        for r in results.values():
            ws.append([
                r["id"], r["name"], r["ip"],
                r["type"], r["serial"], r["pages"],
                r["uzemelteto"], r["cim"]
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        headers={
            "Content-Disposition": "attachment; filename=Teljes_tablazat.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )


# -------------------------
# ADD PRINTER
# -------------------------
@app.route("/add_printer", methods=["POST"])
def add_printer():
    try:
        data = request.get_json()

        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute("""
            INSERT INTO nyomtatok 
            (azonosito, gep_helye, ip, tipus, gyari_szam, uzemelteto, cim, tablazat)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data.get("azonosito"),
            data.get("gep_helye"),
            data.get("ip"),
            data.get("tipus"),
            data.get("gyari_szam"),
            data.get("uzemelteto"),
            data.get("cim"),
            data.get("csoport")
        ))

        db.commit()
        cursor.close()
        db.close()
        invalidate_cache()
        return {"success": True}

    except Exception as e:
        return {"success": False, "error": str(e)}

@app.route("/get_printers")
def get_printers():
    if results:
        uj_results = []

        for id_, adat in results.items():
            uj_results.append({
                "azonosito": adat.get("id"),
                "cim": adat.get("cim"),
                "gep_helye": adat.get("name"),
                "gyari_szam": adat.get("serial"),
                "ip": adat.get("ip"),
                "oldalszam": adat.get("pages"),
                "tablazat": adat.get("tablazat"),
                "tipus": adat.get("type"),
                "rogzitve": adat.get("rogzitve"),
                "uzemelteto": adat.get("uzemelteto"),
                "status": adat.get("status", "idle")
            })
        return jsonify(uj_results)

    else:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT * FROM nyomtatok")
        rows = cursor.fetchall()

        for r in rows:
            r["oldalszam"] = r.get("oldalszam") or "N/A"
            r["gyari_szam"] = r.get("gyari_szam") or "N/A"
            r["tipus"] = r.get("tipus") or "N/A"
            r["ip"] = r.get("ip") or "N/A"
            r["status"] = "idle"
            r["rogzitve"] = r.get("updated_at")

        cursor.close()
        db.close()
        return jsonify(rows)

# -------------------------
# DELETE PRINTER
# -------------------------
@app.route("/delete_printer/<azonosito>", methods=["DELETE"])
def delete_printer(azonosito):
    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("DELETE FROM nyomtatok WHERE azonosito=%s", (azonosito,))
    db.commit()

    cursor.close()
    db.close()
    invalidate_cache()
    return {"success": True}

# -------------------------
# UPDATE PRINTER
# -------------------------
@app.route("/update_printer", methods=["POST"])
def update_printer():
    data = request.json

    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        UPDATE nyomtatok SET
            gep_helye=%s,
            ip=%s,
            tipus=%s,
            gyari_szam=%s,
            uzemelteto=%s,
            cim=%s
        WHERE azonosito=%s
    """, (
        data["gep_helye"],
        data["ip"],
        data["tipus"],
        data["gyari_szam"],
        data["uzemelteto"],
        data["cim"],
        data["azonosito"]
    ))

    db.commit()
    cursor.close()
    db.close()
    invalidate_cache()
    return {"success": True}



@app.route("/update_printer_tablazat", methods=["POST"])
def update_printer_tablazat():
    data = request.json

    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("""
        UPDATE nyomtatok SET
            tablazat=%s
        WHERE azonosito=%s
    """, (
        data["tablazat"],
        data["azonosito"]
    ))

    db.commit()
    cursor.close()
    db.close()
    return {"success": True}

# -------------------------
# SNAPSHOT ROUTE
# -------------------------
@app.route("/save_monthly", methods=["POST"])
def save_monthly():
    try:
        save_monthly_snapshot()
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.route("/get_full_diff")
def getfulldiff():
    honapok = {
        "January": "Január",
        "February": "Február",
        "March": "Március",
        "April": "Április",
        "May": "Május",
        "June": "Június",
        "July": "Július",
        "August": "Augusztus",
        "September": "Szeptember",
        "October": "Október",
        "November": "November",
        "December": "December"
    }
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    

    cursor.execute("""
        SELECT 
            curr.nyomtato_id, 
            curr.uzemelteto, 
            curr.cim, 
            curr.honap_nev, 
            curr.oldalszam AS aktualis, 
            prev.oldalszam AS elozo, 
            (curr.oldalszam - prev.oldalszam) AS kulonbseg 
        FROM 
            nyomtato_havi_allas curr 
            LEFT JOIN nyomtato_havi_allas prev ON curr.nyomtato_id = prev.nyomtato_id 
            AND prev.datum = DATE_SUB(curr.datum, INTERVAL 1 MONTH) 
        WHERE 
            curr.honap_nev LIKE MONTHNAME(
                CURDATE()
            ) 
        ORDER BY 
            `uzemelteto`, 
            `cim`;
    """)
    rows = cursor.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"
    ws.append(["Nyomtató ID","Üzemeltető","Cím","Aktuális hónap","Aktuális nyomatszám","Előző nyomatszám","Különbség"])
    for r in rows:
        ws.append([r["nyomtato_id"], r["uzemelteto"], r["cim"], honapok.get(r["honap_nev"]), r["aktualis"], r["elozo"], r["kulonbseg"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=Tárgy havi nyomatszámok.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "X-Content-Type-Options": "nosniff"  # Chrome biztonságosabbnak látja
    }

    return Response(output, headers=headers)

@app.route('/api/cim')
def cim():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM cim")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)  # [{id: 1, cim: "4200 Hajdúszoboszló ..."}, ...]

@app.route('/api/uzemelteto')
def uzemelteto():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM uzemelteto")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)  # [{id: 1, uzemelteto: "Kerekes Kft."}, ...]

@app.route('/api/add_cim', methods=['POST'])
def add_cim():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO cim (cim) VALUES (%s)", (data['value'],))
    conn.commit()
    cursor.close()
    conn.close()
    return {"status": "ok"}


@app.route('/api/add_uzemelteto', methods=['POST'])
def add_uzemelteto():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO uzemelteto (uzemelteto) VALUES (%s)", (data['value'],))
    conn.commit()
    cursor.close()
    conn.close()
    return {"status": "ok"}

@app.route('/api/list_uzemelteto', methods=['GET'])
def list_uzemelteto():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM uzemelteto")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)

@app.route('/api/list_cim', methods=['GET'])
def list_cim():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM cim")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)


@app.route("/api/delete_cim", methods=["POST"])
def delete_cim():
    data = request.get_json()
    id_to_delete = data.get("id")
    if not id_to_delete:
        return jsonify({"success": False, "error": "Hiányzó ID"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM cim WHERE id = %s", (id_to_delete,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/delete_uzemelteto", methods=["POST"])
def delete_uzemelteto():
    data = request.get_json()
    id_to_delete = data.get("id")
    if not id_to_delete:
        return jsonify({"success": False, "error": "Hiányzó ID"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM uzemelteto WHERE id = %s", (id_to_delete,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
@app.route("/api/update_cim", methods=["POST"])
def update_cim():
    data = request.get_json()
    id_to_update = data.get("id")
    new_value = data.get("value")
    
    if not id_to_update or new_value is None:
        return jsonify({"success": False, "error": "Hiányzó adat"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE cim SET cim = %s WHERE id = %s", (new_value, id_to_update))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/update_uzemelteto", methods=["POST"])
def update_uzemelteto():
    data = request.get_json()
    id_to_update = data.get("id")
    new_value = data.get("value")
    
    if not id_to_update or new_value is None:
        return jsonify({"success": False, "error": "Hiányzó adat"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE uzemelteto SET uzemelteto = %s WHERE id = %s", (new_value, id_to_update))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/get_relations")
def get_all_relations():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT uc.id AS relation_id, u.id AS uzemelteto_id, u.uzemelteto,
               c.id AS cim_id, c.cim
        FROM uzemelteto_cim uc
        JOIN uzemelteto u ON uc.uzemelteto_id = u.id
        JOIN cim c ON uc.cim_id = c.id
        ORDER BY u.uzemelteto, c.cim
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)

@app.route("/api/add_relation", methods=["POST"])
def add_relation():
    data = request.json
    uzem_id = data.get("uzemelteto_id")
    cim_id = data.get("cim_id")
    if not uzem_id or not cim_id:
        return jsonify({"success": False, "error": "Hiányzó adat"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT IGNORE INTO uzemelteto_cim (uzemelteto_id, cim_id) VALUES (%s,%s)",
            (uzem_id, cim_id)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/delete_relation", methods=["POST"])
def delete_relation():
    data = request.json
    relation_id = data.get("relation_id")
    if not relation_id:
        return jsonify({"success": False, "error": "Hiányzó adat"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM uzemelteto_cim WHERE id=%s", (relation_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/get_relations_by_uzem/<int:uzem_id>")
def get_relations_by_uzem(uzem_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT uc.id AS relation_id, c.id AS cim_id, c.cim
        FROM uzemelteto_cim uc
        JOIN cim c ON uc.cim_id = c.id
        WHERE uc.uzemelteto_id = %s
        ORDER BY c.cim
    """, (uzem_id,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(rows)


@app.route("/api/get_csoportok")
def get_csoportok():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT id, csoport FROM csoportok")
        rows = cursor.fetchall()

        result = []
        for row in rows:
            result.append({
                "id": row[0],
                "csoport": row[1]
            })

        cursor.close()
        conn.close()

        return jsonify(result)

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/add_csoport", methods=["POST"])
def add_csoport():
    data = request.get_json()
    name = data.get("csoport")

    if not name:
        return jsonify({"success": False, "error": "Hiányzó név"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            "INSERT INTO csoportok (csoport) VALUES (%s)",
            (name,)
        )

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/update_printer_count", methods=["POST"])
def update_printer_count():
    global results
    data = request.get_json()
    printer_id = data.get("printer_id")
    page_count = data.get("page_count")

    try:
        if page_count not in [None, "", "N/A"]:
            page_count_int = int(page_count)
        else:
            return jsonify({"success": False, "error": "Page count must be an integer."}), 500
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"""
            UPDATE nyomtatok
            SET oldalszam = {page_count_int}
            WHERE azonosito LIKE {printer_id}
        """)
        conn.commit()
        cursor.close()
        conn.close()
        invalidate_cache()
        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/update_csoport", methods=["POST"])
def update_csoport():
    data = request.get_json()
    id_ = data.get("id")
    name = data.get("csoport")

    if not id_ or not name:
        return jsonify({"success": False, "error": "Hiányzó adat"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            "UPDATE csoportok SET csoport=%s WHERE id=%s",
            (name, id_)
        )

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/delete_csoport", methods=["POST"])
def delete_csoport():
    data = request.get_json()
    id_ = data.get("id")

    if not id_:
        return jsonify({"success": False, "error": "Hiányzó ID"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM csoportok WHERE id=%s", (id_,))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@login_manager.user_loader
def load_user(user_id):

    db = None
    cursor = None

    try:
        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute(
            """
            SELECT felhasznalonev, jelszo
            FROM felhasznalok
            WHERE felhasznalonev = %s
            """,
            (user_id,)
        )

        result = cursor.fetchone()

        if result is None:
            return None

        return User(
            username=result[0],
            password_hash=result[1]
        )

    except Exception as e:
        print("USER LOADER HIBA:", e)
        return None

    finally:
        if cursor:
            cursor.close()

        if db:
            db.close()


@app.route("/index")
@login_required
def index():

    print("INDEX:", current_user.username)

    return render_template(
        "index.html",
        felhasznalonev=current_user.username
    )


@app.route("/logout")
@login_required
def logout():

    logout_user()

    return redirect(url_for("login"))


init_app_state()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)