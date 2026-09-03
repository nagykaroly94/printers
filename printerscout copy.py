import io
import os
import configparser
import ipaddress
import bcrypt
import queue

from openpyxl import Workbook
from datetime import datetime, timedelta
from modules.printer_snmp import get_printer_data, SNMPError
from modules.db import execute_statement
from flask import Flask, Response, jsonify, render_template, request, redirect, url_for
from flask_login import LoginManager, UserMixin, login_user, logout_user, current_user

config = configparser.ConfigParser()
config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.ini")

def validate_config(config):
    if not config.has_section("mysql"): raise RuntimeError("A config.ini-ben lennie kell [mysql] szakasznak!")
    for key in ("host", "user", "password", "database"):
        if not config["mysql"].get(key): raise RuntimeError(f"A [mysql] szakaszból hiányzik a(z) '{key}' beállítás!")
    if not config.has_section("app"): raise RuntimeError("A config.ini-ben lennie kell [app] szakasznak!")
    if not config["app"].get("secret"): raise RuntimeError("Az [app] szakaszból hiányzik a 'secret' beállítás!")

try:
    config.read(config_path)
    validate_config(config)
except Exception as e: 
    raise RuntimeError(f"Hibás konfiguráció: {e}")

app = Flask(__name__)
app.config["REMEMBER_COOKIE_DURATION"] = timedelta(days=40)  # 40 napos cookie élettartam
app.config["SECRET_KEY"] = config["app"]["secret"]

clients = []

login_manager = LoginManager()
login_manager.init_app(app)

login_manager.login_view = "login"
login_manager.login_message = "A folytatáshoz be kell jelentkezned."
login_manager.login_message_category = "error"


class User(UserMixin):
    def __init__(self, username, password_hash, isadmin=False):
        self.id = username
        self.username = username
        self.password_hash = password_hash
        self.isadmin = isadmin

def db_execute(statement, params=()):
    try: return execute_statement(config, statement, params)
    except Exception as e: notify_clients_error(e)

def send_snmp_status(id, status, count=0):
    pass

#TODO LOCK OUT DB CHANGES DURING
@app.route("/start")
async def start():
    db_nyomtatok = db_execute("SELECT * FROM nyomtatok")
    for nyomtato in db_nyomtatok:
        db_ip = nyomtato["ip"]
        db_azonosito = nyomtato["azonosito"]
        db_tipus = nyomtato["tipus"]
        db_oldalszam = nyomtato["oldalszam"]
        db_gyari_szam = nyomtato["gyari_szam"]

        try: ipaddress.ip_address(db_ip)
        except ValueError: continue
    
        try:
            lekert_tipus, lekert_oldalszam, lekert_gyari_szam = await get_printer_data(db_ip)

            if db_tipus in ("", None) and lekert_tipus not in ("", None): db_execute("UPDATE nyomtatok SET tipus=%s WHERE azonosito=%s", (lekert_tipus, db_azonosito))
            if lekert_oldalszam.isdigit() and str(db_oldalszam).isdigit():
                if int(lekert_oldalszam) > int(db_oldalszam): db_execute("UPDATE nyomtatok SET oldalszam=%s WHERE azonosito=%s",(lekert_oldalszam, db_azonosito))
            if db_gyari_szam in ("", None): db_execute("UPDATE nyomtatok SET gyari_szam=%s WHERE azonosito=%s",(lekert_gyari_szam, db_azonosito))

            send_snmp_status(id, "success", lekert_oldalszam)
        except SNMPError as exc:
            if "timeout" in exc.message: send_snmp_status(id, "timeout")
            else: notify_clients_error(f"SNMP Lekérdezési hiba: {exc}")
        
"""
        # -------------------------
        # PAGE COUNT LOGIKA
        # -------------------------
        try:
            page_count_int = int(page_count) if page_count not in [None, "", "N/A"] else int(r.get("oldalszam") or 0)
        except:
            page_count_int = None

        old_pages = int(r.get("oldalszam") or 0)
        pages_changed = (
            page_count_int is not None and
            old_pages is not None and
            page_count_int != old_pages
        )

        # -------------------------
        # FINAL OUTPUT (FRONTEND)
        # -------------------------
        final_type = printer_type or r.get("tipus")
        final_serial = serial or r.get("gyari_szam")

        data = {
            "id": r["azonosito"],
            "name": r.get("gep_helye") or "N/A",
            "ip": r.get("ip") or "N/A",
            "type": final_type or "N/A",
            "serial": final_serial or "N/A",
            "pages": page_count_int or "N/A",
            "cim": r.get("cim") or "N/A",
            "uzemelteto": r.get("uzemelteto") or "N/A",
            "tablazat": r.get("tablazat"),
            "status": "ok" if success else "error",
            "rogzitve": (r.get("updated_at").isoformat() if isinstance(r.get("updated_at"), datetime) else r.get("updated_at"))
        }

        local_results[r["azonosito"]] = data

        with live_updates_lock:
            live_updates.append(data)

        # -------------------------
        # DB UPDATE LOGIKA
        # -------------------------
        updates = []
        values = []

        # csak ha jött SNMP adat
        if printer_type:
            updates.append("tipus=%s")
            values.append(printer_type)

        if serial:
            updates.append("gyari_szam=%s")
            values.append(serial)

        if pages_changed:
            updates.append("oldalszam=%s")
            values.append(page_count_int)

        if success:
            updates.append("updated_at = %s")
            values.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # csak akkor írunk DB-be, ha van mit frissíteni
        if updates:
            values.append(r["azonosito"])

            cursor.execute(f" UPDATE nyomtatok SET {", ".join(updates)} WHERE azonosito=%s", values)
            db.commit()
"""

@app.before_request
def check_login():
    # Login oldal, CSS, JS, képek stb. szabadon elérhetők, minden más csak bejelentkezve
    if request.endpoint in ("login", "static"): return
    if not current_user.is_authenticated: return redirect(url_for("login"))

@login_manager.user_loader
def load_user(user_id):
    result = db_execute("SELECT felhasznalonev, jelszo, isadmin FROM felhasznalok WHERE felhasznalonev = %s", user_id)
    if not result: return None
    user = result[0]
    return User(username=user["felhasznalonev"], password_hash=user["jelszo"], isadmin=user["isadmin"])

@app.route("/")
def index(): return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated: return redirect(url_for("index"))

    if request.method == "POST":

        username = request.form.get("username")
        password = request.form.get("password")

        try:
            result = db_execute("SELECT felhasznalonev, jelszo, isadmin FROM felhasznalok WHERE felhasznalonev = %s", username)
            if result:
                user_data = result[0]
                stored_password = user_data["jelszo"]
                isadmin = user_data["isadmin"]

                if bcrypt.checkpw(
                    password.encode("utf-8"),
                    stored_password.encode("utf-8")
                ):

                    user = User(username, stored_password, isadmin)
                    login_user(user, remember=request.form.get("remember") == "on")

                    return redirect(url_for("index"))

            return render_template("login.html", hiba="Helytelen felhasználónév vagy jelszó.")

        except Exception: return render_template("login.html", hiba="Hiba történt a bejelentkezés során.")
    return render_template("login.html")

@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for("login"))

@app.route("/users")
def users(): return render_template("users.html")

@app.route("/events")
def events():
    q = queue.Queue()
    clients.append(q)

    def generate():
        try:
            while True:
                message = q.get()
                yield f"data: {message}\n\n"
        finally:
            clients.remove(q)

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )

def notify_clients_error(message):
    notify_clients("HIBA! " + message)

def notify_clients(message):
    for q in clients:
        q.put(message)

@app.route("/printer_pages/<path:table_name>.xlsx")
def download_table_xlsx(table_name):
    rows = db_execute("SELECT * FROM nyomtatok WHERE tablazat = %s", table_name) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"

    ws.append(["Azonosító", "Hely", "IP cím", "Típus", "Sorozatszám", "Oldalszám"])

    for r in rows:
        ws.append([
            r["azonosito"],
            r["gep_helye"] or "N/A",
            r["ip"] or "N/A",
            r["tipus"] or "N/A",
            r["gyari_szam"] or "N/A",
            r["uzemelteto"] or "N/A",
            r["cim"] or "N/A",
            r["oldalszam"] or "N/A"
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": f"attachment; filename={table_name}.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

    return Response(output, headers=headers)

# -------------------------
# EXCEL EXPORT (ALL)
# -------------------------
@app.route("/printer_pages.xlsx")
def download_all_xlsx():
    rows = db_execute("SELECT * FROM nyomtatok") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"

    ws.append(["Azonosító","Hely","IP","Típus","Sorozatszám","Oldalszám","Üzemeltető","Cím"])

    for r in rows:
        ws.append([
            r["azonosito"],
            r["gep_helye"] or "N/A",
            r["ip"] or "N/A",
            r["tipus"] or "N/A",
            r["gyari_szam"] or "N/A",
            r["uzemelteto"] or "N/A",
            r["cim"] or "N/A",
            r["oldalszam"] or "N/A"
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        headers={
            "Content-Disposition": "attachment; filename=Teljes_tablazat.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )

@app.route("/add_printer", methods=["POST"])
def add_printer():
    data = request.get_json()

    if any(data.get(key) is None for key in ("azonosito", "gep_helye", "ip", "tipus", "gyari_szam", "uzemelteto", "cim", "csoport")):
        notify_clients_error("/add_printer - Hiányzó adat!")
        return

    if db_execute("INSERT INTO nyomtatok (azonosito, gep_helye, ip, tipus, gyari_szam, uzemelteto, cim, tablazat) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", (
        data.get("azonosito"),
        data.get("gep_helye"),
        data.get("ip"),
        data.get("tipus"),
        data.get("gyari_szam"),
        data.get("uzemelteto"),
        data.get("cim"),
        data.get("csoport")
    )) == 1: notify_clients("Nyomtató hozzáadva.")

@app.route("/get_printers")
def get_printers(): return jsonify(db_execute("SELECT * FROM nyomtatok") or [])

@app.route("/delete_printer/<azonosito>", methods=["DELETE"])
def delete_printer(azonosito):
    if not azonosito:
        notify_clients_error("/delete_printer - Hiányzó azonosító!")
        return

    if db_execute("DELETE FROM nyomtatok WHERE azonosito=%s", azonosito) == 1: notify_clients("Nyomtató törölve.")

@app.route("/update_printer", methods=["POST"])
def update_printer():
    data = request.json

    if any(data.get(key) is None for key in ( "gep_helye", "ip", "tipus", "gyari_szam", "uzemelteto", "cim", "azonosito")):
        notify_clients_error("/update_printer - Hiányzó adat!")
        return

    if db_execute("""
        UPDATE nyomtatok SET
            gep_helye=%s,
            ip=%s,
            tipus=%s,
            gyari_szam=%s,
            uzemelteto=%s,
            cim=%s
        WHERE azonosito=%s
    """, (
        data["gep_helye"],
        data["ip"],
        data["tipus"],
        data["gyari_szam"],
        data["uzemelteto"],
        data["cim"],
        data["azonosito"]
    )) == 1:
        notify_clients("Nyomtató módosítva.")

@app.route("/update_printer_tablazat", methods=["POST"])
def update_printer_tablazat():
    data = request.json
    tablazat = data.get("tablazat")
    azonosito = data.get("azonosito")

    if tablazat is None or azonosito is None:
        notify_clients_error("/update_printer_tablazat - Hiányzó adat!")
        return

    if db_execute("UPDATE nyomtatok SET tablazat=%s WHERE azonosito=%s", (tablazat, azonosito)) == 1: notify_clients("Nyomtató táblázata módosítva.")

@app.route("/update_printer_tablazat", methods=["POST"])
def update_printer_tablazat():
    data = request.json
    tablazat = data.get("tablazat")
    azonosito = data.get("azonosito")

    if tablazat is None or azonosito is None:
        notify_clients_error("/update_printer_tablazat - Hiányzó adat!")
        return

    if db_execute("UPDATE nyomtatok SET tablazat=%s WHERE azonosito=%s", (tablazat, azonosito)) == 1: notify_clients("Nyomtató táblázata módosítva.")

@app.route("/save_monthly", methods=["POST"])
def save_monthly():
    if db_execute("""
        INSERT INTO nyomtato_havi_allas (nyomtato_id, uzemelteto, cim, datum, oldalszam)
        SELECT azonosito, uzemelteto, cim, DATE_FORMAT(CURDATE(), '%Y-%m-01'), oldalszam
        FROM nyomtatok
        ON DUPLICATE KEY UPDATE
            oldalszam = VALUES(oldalszam),
            rogzitve = NOW(),
            uzemelteto = VALUES(uzemelteto),
            cim = VALUES(cim);
    """) is not None: notify_clients("Havi állások mentve.")

@app.route("/get_full_diff")
def getfulldiff():
    honapok = {
        "January": "Január",
        "February": "Február",
        "March": "Március",
        "April": "Április",
        "May": "Május",
        "June": "Június",
        "July": "Július",
        "August": "Augusztus",
        "September": "Szeptember",
        "October": "Október",
        "November": "November",
        "December": "December"
    }

    rows = db_execute("""
        SELECT 
            curr.nyomtato_id, 
            curr.uzemelteto, 
            curr.cim, 
            curr.honap_nev, 
            curr.oldalszam AS aktualis, 
            prev.oldalszam AS elozo, 
            (curr.oldalszam - prev.oldalszam) AS kulonbseg 
        FROM 
            nyomtato_havi_allas curr 
            LEFT JOIN nyomtato_havi_allas prev ON curr.nyomtato_id = prev.nyomtato_id 
            AND prev.datum = DATE_SUB(curr.datum, INTERVAL 1 MONTH) 
        WHERE 
            curr.honap_nev LIKE MONTHNAME(
                CURDATE()
            ) 
        ORDER BY 
            `uzemelteto`, 
            `cim`;
    """)

    if rows is None: return Response(status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nyomtatók"
    ws.append(["Nyomtató ID","Üzemeltető","Cím","Aktuális hónap","Aktuális nyomatszám","Előző nyomatszám","Különbség"])
    for r in rows:
        ws.append([r["nyomtato_id"], r["uzemelteto"], r["cim"], honapok.get(r["honap_nev"]), r["aktualis"], r["elozo"], r["kulonbseg"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=Tárgy havi nyomatszámok.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "X-Content-Type-Options": "nosniff"  # Chrome biztonságosabbnak látja
    }

    return Response(output, headers=headers)

@app.route('/api/cim')
def cim(): return jsonify(db_execute("SELECT * FROM cim"))

@app.route('/api/uzemelteto')
def uzemelteto(): return jsonify(db_execute("SELECT * FROM uzemelteto"))

@app.route("/api/add_cim", methods=["POST"])
def add_cim():
    data = request.json
    value = data.get("value")

    if value is None:
        notify_clients_error("/api/add_cim - Hiányzó adat!")
        return

    if db_execute("INSERT INTO cim (cim) VALUES (%s)", value) == 1: notify_clients("Cím hozzáadva.")

@app.route("/api/add_uzemelteto", methods=["POST"])
def add_uzemelteto():
    data = request.json
    value = data.get("value")

    if value is None:
        notify_clients_error("/api/add_uzemelteto - Hiányzó adat!")
        return

    if db_execute("INSERT INTO uzemelteto (uzemelteto) VALUES (%s)", value) == 1: notify_clients("Üzemeltető hozzáadva.")

@app.route('/api/list_uzemelteto', methods=['GET'])
def list_uzemelteto(): return jsonify(db_execute("SELECT * FROM uzemelteto"))

@app.route('/api/list_cim', methods=['GET'])
def list_cim(): return jsonify(db_execute("SELECT * FROM cim"))

@app.route("/api/delete_cim", methods=["POST"])
def delete_cim():
    data = request.get_json()
    id_to_delete = data.get("id")

    if not id_to_delete:
        notify_clients_error("/api/delete_cim - Hiányzó ID!")
        return

    if db_execute("DELETE FROM cim WHERE id = %s", id_to_delete) == 1: notify_clients("Cím törölve.")

@app.route("/api/delete_uzemelteto", methods=["POST"])
def delete_uzemelteto():
    data = request.get_json()

    id_to_delete = data.get("id")
    if not id_to_delete:
        notify_clients_error("/api/delete_uzemelteto - Hiányzó ID!")
        return

    if db_execute("DELETE FROM uzemelteto WHERE id = %s", id_to_delete) == 1: notify_clients("Üzemeltető törölve.")
    
@app.route("/api/update_cim", methods=["POST"])
def update_cim():
    data = request.get_json()
    id_to_update = data.get("id")
    new_value = data.get("value")
    
    if not id_to_update or new_value is None:
        notify_clients_error("/api/update_cim - Hiányzó adat")
        return

    if db_execute("UPDATE cim SET cim = %s WHERE id = %s", (new_value, id_to_update)) == 1: notify_clients("Cím módosítva.")

@app.route("/api/update_uzemelteto", methods=["POST"])
def update_uzemelteto():
    data = request.get_json()
    id_to_update = data.get("id")
    new_value = data.get("value")
    
    if not id_to_update or new_value is None:
        notify_clients_error("/api/update_uzemelteto - Hiányzó adat")
        return

    if db_execute("UPDATE uzemelteto SET uzemelteto = %s WHERE id = %s", (new_value, id_to_update)) == 1: notify_clients("Üzemeltető módosítva.")

@app.route("/api/get_relations")
def get_all_relations():
    return jsonify(db_execute("""
        SELECT uc.id AS relation_id, u.id AS uzemelteto_id, u.uzemelteto,
               c.id AS cim_id, c.cim
        FROM uzemelteto_cim uc
        JOIN uzemelteto u ON uc.uzemelteto_id = u.id
        JOIN cim c ON uc.cim_id = c.id
        ORDER BY u.uzemelteto, c.cim
    """))

@app.route("/api/add_relation", methods=["POST"])
def add_relation():
    data = request.json
    uzem_id = data.get("uzemelteto_id")
    cim_id = data.get("cim_id")
    if not uzem_id or not cim_id: notify_clients_error("/api/add_relation - Hiányzó adat")
    elif db_execute("INSERT IGNORE INTO uzemelteto_cim (uzemelteto_id, cim_id) VALUES (%s,%s)", (uzem_id, cim_id)) == 1: notify_clients("Kapcsolat sikeresen hozzáadva!")

@app.route("/api/delete_relation", methods=["POST"])
def delete_relation():
    data = request.json
    relation_id = data.get("relation_id")

    if not relation_id: notify_clients_error("/api/delete_relation - Hiányzó relation_id!")
    elif db_execute("DELETE FROM uzemelteto_cim WHERE id=%s", relation_id) == 1: notify_clients("Kapcsolat sikeresen törölve!")

@app.route("/api/get_relations_by_uzem/<int:uzem_id>")
def get_relations_by_uzem(uzem_id):
    if not uzem_id:
        notify_clients_error("/api/get_relations_by_uzem - Hiányzó üzemeltető ID!")
    else: return jsonify(db_execute("""
            SELECT uc.id AS relation_id, c.id AS cim_id, c.cim
            FROM uzemelteto_cim uc
            JOIN cim c ON uc.cim_id = c.id
            WHERE uc.uzemelteto_id = %s
            ORDER BY c.cim
        """, uzem_id))

@app.route("/api/get_csoportok")
def get_csoportok(): return jsonify(db_execute("SELECT id, csoport FROM csoportok"))

@app.route("/api/add_csoport", methods=["POST"])
def add_csoport():
    data = request.get_json()
    name = data.get("csoport")
    if not name: notify_clients_error("Hiányzó csoportnév!")
    elif db_execute("INSERT INTO csoportok (csoport) VALUES (%s)", name) == 1: notify_clients("Csoport sikeresen hozzáadva!")

@app.route("/api/update_printer_count", methods=["POST"])
def update_printer_count():
    data = request.get_json()
    printer_id = data.get("printer_id")
    page_count = data.get("page_count")

    if not printer_id:
        notify_clients_error("/api/update_printer_count - Hiányzó nyomtató ID!")
        return

    if page_count in [None, "", "N/A"]:
        notify_clients_error("Kötelező oldalszámot megadni!")
        return

    try: page_count = int(page_count)
    except (ValueError, TypeError):
        notify_clients_error("Az oldalszámnak számnak kell lennie!")
        return

    if db_execute("UPDATE nyomtatok SET oldalszam=%s WHERE azonosito=%s",(page_count, printer_id)) == 1: notify_clients("Oldalszám módosítva.")

@app.route("/api/update_csoport", methods=["POST"])
def update_csoport():
    data = request.get_json()
    id_ = data.get("id")
    name = data.get("csoport")
    if id_ and name:
        if db_execute("UPDATE csoportok SET csoport=%s WHERE id=%s", (name, id_)) == 1: notify_clients("Csoport módosítva.")
    else: notify_clients_error("/api/update_csoport - Hiányzó adat!")

@app.route("/api/delete_csoport", methods=["POST"])
def delete_csoport():
    data = request.get_json()
    id_ = data.get("id")
    if not id_: notify_clients_error("/api/delete_csoport - Hiányzó ID!")
    elif db_execute("DELETE FROM csoportok WHERE id=%s", id_) == 1:  notify_clients("Sikeres törlés!")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)